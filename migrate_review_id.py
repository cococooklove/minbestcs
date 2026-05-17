"""
이미 다운로드된 엑셀에서 review_id를 추출해 reviews.json에 채운다.
scraper.py 재실행 없이 기존 reviews.json을 그대로 업데이트.

실행:
    python3 migrate_review_id.py [엑셀_경로]
기본 엑셀: data/downloads/ 중 가장 최신.
"""
import json
import os
import sys
import re
import glob
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")


REVIEWS_FILE = "data/reviews.json"


def _norm_date(s):
    s = str(s or "")
    m = re.match(r'(\d{4})\.(\d{2})\.(\d{2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return s


def main():
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        cands = sorted(glob.glob("data/downloads/*.xlsx"), key=os.path.getmtime, reverse=True)
        if not cands:
            print("data/downloads/ 에 엑셀 없음")
            return 1
        excel_path = cands[0]

    print(f"엑셀: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name, exclude=None):
        for i, h in enumerate(headers):
            if h and name in str(h):
                if exclude and exclude in str(h):
                    continue
                return i + 1
        return None

    c_rid = col("리뷰글번호", exclude="관련")
    c_date = col("리뷰등록일")
    c_content = col("리뷰상세내용", exclude="관련")
    c_reviewer = col("등록자")
    if not all([c_rid, c_date, c_content, c_reviewer]):
        print(f"필수 컬럼 누락: rid={c_rid}, date={c_date}, content={c_content}, reviewer={c_reviewer}")
        return 1
    print(f"컬럼: review_id={c_rid}, date={c_date}, content={c_content}, reviewer={c_reviewer}")

    # 엑셀 키 → review_id
    mapping = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, c_rid).value
        if not rid:
            continue
        key = (
            str(ws.cell(r, c_content).value or "").strip(),
            _norm_date(ws.cell(r, c_date).value),
            str(ws.cell(r, c_reviewer).value or "").strip(),
        )
        mapping[key] = str(rid)
    print(f"엑셀에서 review_id {len(mapping)}건 추출")

    # reviews.json 업데이트
    with open(REVIEWS_FILE, encoding="utf-8") as f:
        reviews = json.load(f)
    matched = 0
    already = 0
    missing = 0
    for r in reviews:
        if r.get("review_id"):
            already += 1
            continue
        key = (
            r.get("content", "").strip(),
            r.get("date", "").strip(),
            r.get("reviewer", "").strip(),
        )
        if key in mapping:
            r["review_id"] = mapping[key]
            matched += 1
        else:
            r["review_id"] = ""
            missing += 1

    with open(REVIEWS_FILE, "w", encoding="utf-8") as f:
        json.dump(reviews, f, ensure_ascii=False, indent=2)
    print(f"\n=== 결과 ===")
    print(f"  전체:     {len(reviews)}")
    print(f"  이미 있음: {already}")
    print(f"  매칭됨:   {matched}")
    print(f"  매칭 실패: {missing}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
